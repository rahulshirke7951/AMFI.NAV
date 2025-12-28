import os
import json
import pandas as pd
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE_DIR, "scheme_rules.json")) as f:
    RULES = json.load(f)

# --------------------------------------------------
# Helpers
# --------------------------------------------------

def normalize(s):
    return " ".join(str(s).upper().split())

def clean_text(s):
    for ch in ["-", "–", "—"]:
        s = s.replace(ch, " ")
    return normalize(s)

def extract_base_scheme(name):
    s = clean_text(name)
    for t in RULES["base_scheme_remove_terms"]:
        s = s.replace(t, "")
    return normalize(s)

def exclusion_reason(name):
    for k in RULES["exclusion_rules"]["contains_any"]:
        if k in name.upper():
            return f"Excluded by rule: contains {k}"
    return None

def select_variant(grp):
    for rule in RULES["selection_rules"]["priority_ladder"]:
        m = grp[grp["Mutual Fund Name"].str.upper().apply(lambda x: all(k in x for k in rule))]
        if not m.empty:
            return m.iloc[0], grp.drop(m.index)
    return grp.iloc[0], grp.iloc[1:]

# --------------------------------------------------
# Flatten merged cells
# --------------------------------------------------

def flatten(file, sheet="NAV Data"):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]
    for m in list(ws.merged_cells.ranges):
        v = ws.cell(m.min_row, m.min_col).value
        ws.unmerge_cells(str(m))
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                ws.cell(r, c).value = v
    out = file.replace(".xlsx", "_flat.xlsx")
    wb.save(out)
    return out

# --------------------------------------------------
# Extract one NAV file
# --------------------------------------------------

def extract(file):
    flat = flatten(file)
    raw = pd.read_excel(flat, header=None).dropna(how="all")

    hdr = raw[raw.apply(lambda r: r.astype(str).str.contains("NAV Name", case=False).any(), axis=1)].index[0]
    raw.columns = raw.iloc[hdr]
    df = raw.iloc[hdr + 1:][["NAV Name", "Net Asset Value"]]
    df.columns = ["Mutual Fund Name", "NAV"]

    df["NAV"] = pd.to_numeric(df["NAV"], errors="coerce")
    df = df.dropna()

    total_raw = len(df)

    rule_excl, included = [], []

    for _, r in df.iterrows():
        reason = exclusion_reason(r["Mutual Fund Name"])
        if reason:
            d = r.to_dict()
            d["Reason"] = reason
            rule_excl.append(d)
        else:
            included.append(r)

    df = pd.DataFrame(included)
    df["Base"] = df["Mutual Fund Name"].apply(extract_base_scheme)
    df["Key"] = df["Mutual Fund Name"].apply(normalize)

    final, variant_excl = [], []

    for _, grp in df.groupby("Base"):
        if len(grp) == 1:
            final.append(grp.iloc[0])
        else:
            keep, drop = select_variant(grp)
            final.append(keep)
            for _, d in drop.iterrows():
                e = d.to_dict()
                e["Reason"] = "Excluded: non-preferred variant"
                variant_excl.append(e)

    final_df = pd.DataFrame(final)
    excl_df = pd.concat([pd.DataFrame(rule_excl), pd.DataFrame(variant_excl)], ignore_index=True)

    return final_df, excl_df, total_raw

# --------------------------------------------------
# Main comparison + reconciliation
# --------------------------------------------------

def run(latest, past):
    l_df, l_rule_excl, l_raw = extract(latest)
    p_df, p_rule_excl, p_raw = extract(past)

    l_df = l_df.rename(columns={"NAV": "Latest NAV"})
    p_df = p_df.rename(columns={"NAV": "Past NAV"})

    merged = pd.merge(l_df, p_df[["Key", "Past NAV"]], on="Key", how="outer", indicator=True)

    # Not comparable
    l_nc = merged[merged["_merge"] == "left_only"].copy()
    p_nc = merged[merged["_merge"] == "right_only"].copy()

    # Comparable
    comp = merged[merged["_merge"] == "both"].copy()

    # Zero NAV
    l_zero = comp[comp["Latest NAV"] == 0].copy()
    p_zero = comp[comp["Past NAV"] == 0].copy()

    comp = comp[(comp["Latest NAV"] != 0) & (comp["Past NAV"] != 0)]

    comp["Change"] = comp["Latest NAV"] - comp["Past NAV"]
    comp["Change %"] = (comp["Change"] / comp["Past NAV"] * 100).round(2)

    nav_comp = comp[["Mutual Fund Name", "Latest NAV", "Past NAV", "Change", "Change %"]]

    # ---------------- Reconciliation ----------------

    rec = pd.DataFrame([
        ["Latest", "Total Raw", l_raw],
        ["Latest", "Included in NAV Comparison", len(nav_comp)],
        ["Latest", "Excluded – Scheme Rules", len(l_rule_excl)],
        ["Latest", "Excluded – Zero NAV", len(l_zero)],
        ["Latest", "Excluded – Not Comparable", len(l_nc)],

        ["Past", "Total Raw", p_raw],
        ["Past", "Included in NAV Comparison", len(nav_comp)],
        ["Past", "Excluded – Scheme Rules", len(p_rule_excl)],
        ["Past", "Excluded – Zero NAV", len(p_zero)],
        ["Past", "Excluded – Not Comparable", len(p_nc)],
    ], columns=["File Type", "Category", "Count"])

    # ---------------- Output ----------------

    with pd.ExcelWriter("NAV_Comparison_Result.xlsx", engine="openpyxl") as w:
        nav_comp.to_excel(w, "NAV Comparison", index=False)
        l_rule_excl.to_excel(w, "Excluded_Scheme_Rules_Latest", index=False)
        p_rule_excl.to_excel(w, "Excluded_Scheme_Rules_Past", index=False)
        l_zero.to_excel(w, "Excluded_Zero_NAV_Latest", index=False)
        p_zero.to_excel(w, "Excluded_Zero_NAV_Past", index=False)
        l_nc.to_excel(w, "Excluded_Not_Comparable_Latest", index=False)
        p_nc.to_excel(w, "Excluded_Not_Comparable_Past", index=False)
        rec.to_excel(w, "Reconciliation", index=False)

# --------------------------------------------------
# Entry
# --------------------------------------------------

if __name__ == "__main__":
    run("data/latest_nav.xlsx", "data/past_nav.xlsx")
